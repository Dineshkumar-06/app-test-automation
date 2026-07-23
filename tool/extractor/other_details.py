"""Resolve dropdown option lists that the SOW keeps in a supplementary
sheet (NITR: "Other Details"; SBI: "Others").

Many dropdown fields don't list their options inline — the Values cell just
says "Refer other details" / "List given in other details", and the real
option list lives in a column of a separate sheet. This module reads that
sheet into named lists and matches a field to its list by header name.

Kept deterministic (no LLM): the field label and the column header are
normalized and matched by alias, containment, or token overlap. An
unresolved reference returns None so the caller can flag it low-confidence
rather than silently emitting an empty dropdown.
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

# Sheets that hold supplementary option lists, by likely name. Discovery is
# case-insensitive and also matches any sheet whose name contains "other".
_SUPPLEMENTARY_SHEET_NAMES = ("Other Details", "Others", "Other details")

# Headers that are not option lists (row indices, blank spacers).
_NON_LIST_HEADERS = {"sr. no.", "sr no", "sl no", "sl no.", "s.no"}

# Known field-label -> column-header aliases for cases token-overlap misses.
_ALIASES = {
    "post applied for": "post names",
    "centre of examination": "centres",
}

# Some option lists aren't their own column — they sit as a sub-block inside
# a busy column (e.g. NITR stacks the exam-centre list under a "Centres"
# label inside the "GST Invoice Calculation" column). When one of these
# anchor words appears in any cell, the contiguous non-empty cells beneath
# it in that column are collected as a named list. Extend as new
# applications reveal more stacked lists.
_SUBLIST_ANCHORS = ("Centres", "Centre")


def _norm(text: str) -> str:
    return re.sub(r"[^a-z0-9 ]", "", str(text).lower()).strip()


def _tokens(text: str) -> set[str]:
    return {t for t in _norm(text).split() if t}


def find_supplementary_sheet(wb) -> str | None:
    for name in _SUPPLEMENTARY_SHEET_NAMES:
        if name in wb.sheetnames:
            return name
    for name in wb.sheetnames:
        if "other" in name.lower():
            return name
    return None


def load_option_lists(path: str | Path, sheet_name: str | None = None) -> dict[str, list[str]]:
    """Return {column header -> ordered unique option values} for every
    labelled column of the supplementary sheet. Returns {} if no such
    sheet exists.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = sheet_name or find_supplementary_sheet(wb)
    if sheet_name is None:
        return {}
    ws = wb[sheet_name]

    lists: dict[str, list[str]] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header is None or not str(header).strip():
            continue
        header = str(header).strip()
        if _norm(header) in _NON_LIST_HEADERS:
            continue
        values: list[str] = []
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row, col).value
            if v is not None and str(v).strip():
                s = str(v).strip()
                if s not in values:
                    values.append(s)
        if values:
            lists[header] = values

    _extract_sublist_blocks(ws, lists)
    return lists


def _extract_sublist_blocks(ws, lists: dict[str, list[str]]) -> None:
    """Add any stacked sub-lists (see _SUBLIST_ANCHORS) to `lists` in place.

    When an anchor word is found in a cell, the contiguous non-empty cells
    directly beneath it in the same column become a list named after the
    anchor. Does not overwrite a real top-level column of the same name.
    """
    anchors = {a.lower() for a in _SUBLIST_ANCHORS}
    for col in range(1, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row, col).value
            if cell is None or str(cell).strip().lower() not in anchors:
                continue
            name = str(cell).strip()
            if name in lists:
                continue
            block: list[str] = []
            for r in range(row + 1, ws.max_row + 1):
                v = ws.cell(r, col).value
                if v is None or not str(v).strip():
                    break
                s = str(v).strip()
                if s.lower() in anchors:  # next sub-block starts
                    break
                if s not in block:
                    block.append(s)
            if block:
                lists[name] = block


def resolve_dropdown(label: str, option_lists: dict[str, list[str]]) -> tuple[str, list[str]] | None:
    """Match a field label to a supplementary option list.

    Returns (matched_header, values) or None if no confident match. Tries,
    in order: explicit alias, exact/containment on normalized text, then
    token-overlap (Jaccard >= 0.5, tie-broken by overlap size).
    """
    if not option_lists:
        return None

    norm_label = _norm(label)
    headers = list(option_lists.keys())
    norm_headers = {h: _norm(h) for h in headers}

    # 1. alias
    for alias_label, alias_header in _ALIASES.items():
        if alias_label in norm_label:
            for h, nh in norm_headers.items():
                if nh == alias_header or alias_header in nh:
                    return h, option_lists[h]

    # 2. exact / containment
    for h, nh in norm_headers.items():
        if nh and (nh == norm_label or nh in norm_label or norm_label in nh):
            return h, option_lists[h]

    # 3. token overlap
    label_tokens = _tokens(label)
    best_header, best_score = None, 0.0
    for h in headers:
        htoks = _tokens(h)
        if not htoks or not label_tokens:
            continue
        jaccard = len(label_tokens & htoks) / len(label_tokens | htoks)
        if jaccard > best_score:
            best_header, best_score = h, jaccard
    if best_header is not None and best_score >= 0.5:
        return best_header, option_lists[best_header]

    return None
