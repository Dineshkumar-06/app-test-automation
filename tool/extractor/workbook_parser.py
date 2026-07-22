"""Deterministic parsing of the clean, columnar parts of the requirement
workbooks (SOW, age criteria, eligibility) into tool.models objects.

No LLM calls here. The SOW's free-text ``Validations`` column is
deliberately NOT read by this module — that goes through the extractor's
LLM step (with confidence tags) in a later stage.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

from tool.models import AgeRule, EligibilityLeaf, EligibilityTree, Field, Source

# ---------------------------------------------------------------------------
# SOW field parsing
# ---------------------------------------------------------------------------

# SOW column headers are stable across applications (see CLAUDE.md); map by
# header text rather than fixed index so column order/width can't break us.
_SOW_HEADER_ALIASES = {
    "sl no": "sl",
    "sl no.": "sl",
    "label name": "label",
    "type": "type",
    "max length": "max_length",
    "mandatory ?": "mandatory",
    "mandatory": "mandatory",
    "input method": "input_method",
    "values": "values",
    "default value": "default",
}

_TYPE_MAP = {
    "alpha": "alpha",
    "alphanumeric": "alphanumeric",
    "alpha numeric": "alphanumeric",
    "numeric": "numeric",
    "date": "date",
    "email": "email",
    "checkbox": "checkbox",
    "check box": "checkbox",
    "varchar": "varchar",
    # No dedicated schema type for these — safe generic fallback bucket.
    # "info"/"textbox"-as-a-type are static/declaration display rows.
    "character": "varchar",
    "characters": "varchar",
    "info": "varchar",
    "textbox": "varchar",
}

_MANDATORY_MAP = {
    "yes": True,
    "no": False,
    "conditional": "conditional",
}


def _normalize_type(raw: str, *, sheet: str, row: int) -> str:
    key = str(raw).strip().lower()
    if key not in _TYPE_MAP:
        raise ValueError(f"{sheet}!row{row}: unrecognized Type value {raw!r}")
    return _TYPE_MAP[key]


def _normalize_mandatory(raw: str, *, sheet: str, row: int):
    key = str(raw).strip().lower()
    if key not in _MANDATORY_MAP:
        raise ValueError(f"{sheet}!row{row}: unrecognized Mandatory value {raw!r}")
    return _MANDATORY_MAP[key]


def _normalize_input_method(raw: str, *, sheet: str, row: int) -> str:
    key = str(raw).strip().lower()
    if "multi" in key:
        return "multiselect"
    if "drop" in key:
        return "dropdown"
    if "radio" in key:
        return "radio"
    if "check" in key:
        return "checkbox"
    if "text" in key or "decimal" in key:
        return "text"
    if "label" in key or "read only" in key:
        return "label"
    raise ValueError(f"{sheet}!row{row}: unrecognized Input Method value {raw!r}")


def _normalize_max_length(raw) -> int | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return int(round(raw))
    match = re.search(r"\d+", str(raw))
    return int(match.group()) if match else None


def _normalize_sl(raw) -> str:
    if isinstance(raw, float) and raw.is_integer():
        return str(int(raw))
    return str(raw).strip()


def _header_row_map(row: tuple) -> dict[str, int]:
    """Map canonical column name -> 0-based index, from a header row."""
    out: dict[str, int] = {}
    for idx, cell in enumerate(row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        if key in _SOW_HEADER_ALIASES:
            out[_SOW_HEADER_ALIASES[key]] = idx
    return out


def parse_sow_sheet(path: str | Path, sheet_name: str) -> tuple[list[Field], list[dict]]:
    """Parse one SOW sheet's clean columns into Field objects.

    Rows with no value in the Mandatory column are treated as non-field rows
    (section headers, free-standing notes, sub-headings) and skipped — every
    real field row in both samples has Yes/No/Conditional there. Rows that
    otherwise look like fields but are missing a Label Name are skipped and
    reported in the second return value, since a field with no label can
    never be resolved on the live page.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    columns = _header_row_map(rows[0])
    required = {"sl", "label", "type", "mandatory", "input_method"}
    missing = required - columns.keys()
    if missing:
        raise ValueError(f"{sheet_name}: SOW header missing columns {sorted(missing)}")

    fields: list[Field] = []
    skipped: list[dict] = []

    for row_num, row in enumerate(rows[1:], start=2):
        def get(name: str):
            idx = columns.get(name)
            return row[idx] if idx is not None else None

        mandatory_raw = get("mandatory")
        if mandatory_raw is None or str(mandatory_raw).strip() == "":
            continue  # section header / note / non-field row

        label = get("label")
        if label is None or str(label).strip() == "":
            skipped.append(
                {"sheet": sheet_name, "row": row_num, "sl": get("sl"), "reason": "missing_label"}
            )
            continue

        values_raw = get("values")
        enum = _parse_enum(values_raw) if values_raw else None

        fields.append(
            Field(
                sl=_normalize_sl(get("sl")),
                label=str(label).strip(),
                type=_normalize_type(get("type"), sheet=sheet_name, row=row_num),
                max_length=_normalize_max_length(get("max_length")),
                mandatory=_normalize_mandatory(mandatory_raw, sheet=sheet_name, row=row_num),
                input_method=_normalize_input_method(
                    get("input_method"), sheet=sheet_name, row=row_num
                ),
                enum=enum,
                default=(str(get("default")).strip() if get("default") is not None else None),
                source=Source(sheet=sheet_name, row=row_num),
            )
        )

    return fields, skipped


# A simple quoted/comma-separated list, e.g. "('SC','ST','OBC','EWS','UR')" or
# "Male / Female / Others". Free-form prose (the common case) is left alone —
# it belongs to the Validations-column LLM step, not here.
_QUOTED_LIST_RE = re.compile(r"^\s*\(?\s*'[^']+'(?:\s*,\s*'[^']+')+\s*\)?\s*$")
# No newlines (a real enum list is one line) and each slash-separated token
# is short (<=3 words) — excludes multi-sentence prose that merely happens
# to contain a "/".
_SLASH_LIST_RE = re.compile(r"^[^\n/]+(?:/[^\n/]+){1,6}$")
_TOKEN_RE = re.compile(r"^\s*(?:\w+\s*){1,3}\s*$")
_LEAD_IN_RE = re.compile(
    r"^\s*(?:values?\s+are|value\s+as|select\s+any\s+one\s+of\s+the\s+values?\s+from)\s*[:\-]?\s*",
    re.IGNORECASE,
)


def _parse_enum(raw: str) -> list[str] | None:
    text = _LEAD_IN_RE.sub("", str(raw).strip())
    if _QUOTED_LIST_RE.match(text):
        return [v.strip() for v in re.findall(r"'([^']+)'", text)]
    if _SLASH_LIST_RE.match(text):
        tokens = [v.strip() for v in text.split("/")]
        if all(_TOKEN_RE.match(t) for t in tokens):
            return tokens
    return None


# ---------------------------------------------------------------------------
# Age criteria parsing
# ---------------------------------------------------------------------------

_CATEGORY_TOKENS = ("UR/EWS", "UR", "EWS", "OBC", "SC", "ST")


def _to_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _find_label_value(rows: list[tuple], start: int, end: int, label: str, offset: int):
    """Scan rows[start:end] for a cell equal to label; return the cell `offset`
    positions to its right. The label's column position varies with how many
    category columns the sheet has, so it is located by content, not index.
    """
    for row in rows[start:end]:
        for idx, cell in enumerate(row):
            if cell is not None and str(cell).strip() == label:
                target = idx + offset
                return row[target] if target < len(row) else None
    return None


def _categories_from_group_label(label: str) -> list[str]:
    text = label.lower()
    if "all category" in text:
        return ["UR", "EWS", "OBC", "SC", "ST"]
    return [tok for tok in ("UR", "EWS", "OBC", "SC", "ST") if re.search(rf"\b{tok}\b", label)]


def _parse_relaxation_grid(rows: list[tuple], header_idx: int):
    """Parse the 'Relaxation/Category' grid starting at rows[header_idx].

    Returns (categories, {row_label: {category: date}}, next_row_index).
    """
    header_row = rows[header_idx]
    section_starts = [i for i, v in enumerate(header_row) if v is not None and i >= 1]
    n = section_starts[1] - section_starts[0]  # categories per metric block
    cat_row = rows[header_idx + 1]
    categories = [str(c).strip() for c in cat_row[1 : 1 + n]]

    # Column layout: [label][relax_years x n][max_years x n][max_dates x n]
    dates_start = 1 + 2 * n
    by_label: dict[str, dict[str, date]] = {}
    i = header_idx + 2
    while i < len(rows):
        row = rows[i]
        name = row[0]
        if name is None or str(name).strip() == "":
            break
        dates = row[dates_start : dates_start + n]
        by_label[str(name).strip()] = {
            cat: d for cat, val in zip(categories, dates) if (d := _to_date(val)) is not None
        }
        i += 1
    return categories, by_label, i


def parse_age_criteria(path: str | Path) -> tuple[dict[str, AgeRule], dict[str, list[str]]]:
    """Parse every sheet of an age-criteria workbook into AgeRule objects.

    Handles the plain single-block layout (one post-group, one relaxation
    grid) and the state-group-split layout (SBI sample: multiple named
    "Vacancy for/only for ..." blocks under one post). In the split case the
    age dict is keyed ``"<post> [<state group>]"`` since the tree bounds
    genuinely differ per group; category_by_post is keyed the same way.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    age: dict[str, AgeRule] = {}
    category_by_post: dict[str, list[str]] = {}

    for sheet_name in wb.sheetnames:
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        if not rows:
            continue

        min_dob = _to_date(_find_label_value(rows, 0, len(rows), "Minimum years", 3))

        posts: list[str] = []
        i = 1
        while i < len(rows) and rows[i][0] is not None:
            posts.append(str(rows[i][0]).strip())
            i += 1
        while i < len(rows) and (rows[i][0] is None or str(rows[i][0]).strip() == ""):
            i += 1

        groups: list[tuple[str | None, list[str]]] = []
        if i < len(rows) and rows[i][0] is not None and str(rows[i][0]).strip().lower().startswith("vacancy"):
            while i < len(rows):
                if rows[i][0] is None or str(rows[i][0]).strip() == "":
                    i += 1
                    continue
                if not str(rows[i][0]).strip().lower().startswith("vacancy"):
                    break
                group_label = str(rows[i][0]).strip()
                i += 1  # 'State' label row
                i += 1
                state_list = [s.strip() for s in str(rows[i][0]).split("/")] if rows[i][0] else []
                i += 1
                while rows[i][0] is None or str(rows[i][0]).strip() == "":
                    i += 1
                categories, by_label, i = _parse_relaxation_grid(rows, i)
                cats = _categories_from_group_label(group_label) or categories
                groups.append((group_label, cats))
                for post in posts:
                    key = f"{post} [{group_label}]"
                    age[key] = AgeRule(
                        min_dob=min_dob,
                        max_dob_by_category=by_label.get("General", {}),
                        relaxations={k: v for k, v in by_label.items() if k != "General"} or None,
                    )
                    category_by_post[key] = cats
                while i < len(rows) and (rows[i][0] is None or str(rows[i][0]).strip() == ""):
                    i += 1
        else:
            categories, by_label, _ = _parse_relaxation_grid(rows, i)
            for post in posts:
                age[post] = AgeRule(
                    min_dob=min_dob,
                    max_dob_by_category=by_label.get("General", {}),
                    relaxations={k: v for k, v in by_label.items() if k != "General"} or None,
                )
                category_by_post[post] = categories

    return age, category_by_post


# ---------------------------------------------------------------------------
# Eligibility parsing
# ---------------------------------------------------------------------------

_ELIGIBILITY_HEADER_ALIASES = {
    "post": "post",
    "post name": "post",
    "field": "qualification",
    "degree": "stream",
    "degree/subject / stream": "stream",
    "percentage of marks": "pct",
    "class/grade": "class",
    "class / grade": "class",
    "work experience": "work_experience",
}

_PCT_RE = re.compile(r"(\d+(?:\.\d+)?)")


def _parse_pct(raw) -> tuple[float, bool | None]:
    text = str(raw).strip()
    match = _PCT_RE.search(text)
    pct = float(match.group(1)) if match else 0.0
    strict = ">" in text and "=" not in text
    return pct, (strict or None)


def _parse_class(raw) -> str:
    text = str(raw).strip()
    return "Any" if text.lower() in ("any class", "any", "") else text


def _make_leaf(row: tuple, cols: dict[str, int]) -> EligibilityLeaf | None:
    def get(name: str):
        idx = cols.get(name)
        return row[idx] if idx is not None else None

    qualification = get("qualification")
    if qualification is None or str(qualification).strip() == "":
        return None
    qualification = str(qualification).strip()

    stream_raw = get("stream")
    pct_raw = get("pct")
    if stream_raw is None and pct_raw is None:
        # descriptive/free-text requirement row, not a qualification leaf
        return EligibilityLeaf(qualification=qualification, stream=[], min_pct=0, class_="Any")

    stream = [s.strip() for s in str(stream_raw).split("/")] if stream_raw else ["Any"]
    min_pct, strict = _parse_pct(pct_raw) if pct_raw is not None else (0.0, None)
    work_experience = get("work_experience")
    work_experience = (
        str(work_experience).strip()
        if work_experience is not None and str(work_experience).strip() not in ("", "-")
        else None
    )
    return EligibilityLeaf(
        qualification=qualification,
        stream=stream,
        min_pct=min_pct,
        class_=_parse_class(get("class")),
        strict=strict,
        work_experience=work_experience,
    )


def _build_tree(leaves: list[EligibilityLeaf], ops: list[str]) -> EligibilityTree:
    """Group a flat leaf sequence by its AND/OR operators.

    AND binds tighter than OR (no parentheses appear in the source): split
    on OR into groups, AND-join within each group, then OR across groups.
    Always produces a uniform {op: OR, children: [{op: AND, children: [...]}]}
    shape so the engine has one tree shape to evaluate regardless of how
    many groups there are.
    """
    groups: list[list[EligibilityLeaf]] = [[]]
    groups[0].append(leaves[0])
    for op, leaf in zip(ops, leaves[1:]):
        if op == "OR":
            groups.append([leaf])
        else:  # AND
            groups[-1].append(leaf)
    return EligibilityTree(
        op="OR", children=[EligibilityTree(op="AND", children=g) for g in groups]
    )


def parse_eligibility(path: str | Path, sheet_name: str | None = None) -> dict[str, EligibilityTree]:
    """Parse an eligibility sheet into one EligibilityTree per post.

    Column positions vary between applications (see samples/), so columns
    are located by header text rather than fixed index. A post's block runs
    from its Post-Name row until the next fully blank row.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    cols: dict[str, int] = {}
    for idx, row in enumerate(rows):
        candidate = {}
        for i, cell in enumerate(row):
            if cell is None:
                continue
            key = str(cell).strip().lower()
            if key in _ELIGIBILITY_HEADER_ALIASES:
                candidate[_ELIGIBILITY_HEADER_ALIASES[key]] = i
        if "post" in candidate and "qualification" in candidate:
            header_idx = idx
            cols = candidate
            break
    if header_idx is None:
        raise ValueError(f"{path}: could not locate eligibility header row")

    post_col = cols["post"]
    trees: dict[str, EligibilityTree] = {}
    current_post: str | None = None
    leaves: list[EligibilityLeaf] = []
    ops: list[str] = []

    def flush():
        if current_post is not None and leaves:
            trees[current_post] = _build_tree(leaves, ops)

    for row in rows[header_idx + 1 :]:
        if all(v is None for v in row):
            flush()
            current_post = None
            leaves, ops = [], []
            continue

        post_val = row[post_col]
        if post_val is not None and str(post_val).strip() != "":
            flush()
            current_post = str(post_val).strip()
            leaves, ops = [], []

        qual_val = row[cols["qualification"]] if cols.get("qualification") is not None else None
        if qual_val is None:
            continue
        text = str(qual_val).strip()
        if text in ("AND", "OR"):
            ops.append(text)
            continue
        leaf = _make_leaf(row, cols)
        if leaf is not None:
            leaves.append(leaf)

    flush()
    return trees


# ---------------------------------------------------------------------------
# CLI — dump parsed output for manual review, e.g.:
#   uv run python -m tool.extractor.workbook_parser sow samples/nitrjul26/Revised_1_SOW.xlsx "Basic Details" > out.json
#   uv run python -m tool.extractor.workbook_parser age samples/nitrjul26/Revised_Age_Criteria.xlsx > out.json
#   uv run python -m tool.extractor.workbook_parser eligibility samples/nitrjul26/Revised_Eligibility.xlsx > out.json
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import json
    import sys

    def _dump(obj) -> str:
        return json.dumps(obj, default=str, indent=2)

    mode = sys.argv[1]
    workbook_path = sys.argv[2]
    sheet_arg = sys.argv[3] if len(sys.argv) > 3 else None

    if mode == "sow":
        fields, skipped = parse_sow_sheet(workbook_path, sheet_arg)
        print(_dump({
            "field_count": len(fields),
            "skipped_count": len(skipped),
            "skipped": skipped,
            "fields": [f.model_dump(exclude_none=True) for f in fields],
        }))
    elif mode == "age":
        age, category_by_post = parse_age_criteria(workbook_path)
        print(_dump({
            "post_count": len(age),
            "category_by_post": category_by_post,
            "age": {k: v.model_dump(exclude_none=True) for k, v in age.items()},
        }))
    elif mode == "eligibility":
        trees = parse_eligibility(workbook_path, sheet_arg)
        print(_dump({
            "post_count": len(trees),
            "eligibility": {k: v.model_dump(exclude_none=True) for k, v in trees.items()},
        }))
    else:
        print(f"unknown mode {mode!r}; expected sow|age|eligibility", file=sys.stderr)
        sys.exit(1)
